import openpyxl as xl
wb = xl.load_workbook("pythontry.xlsx")
sheet = wb["Analizler pier marine hotel"]
# cell = sheet["a1"]
#cell = sheet.cell(3,3)
#print(cell.value)
#print(sheet.max_column)
for row in range(5,sheet.max_row +1):
    for column in range(7,sheet.max_column+1):
        cell = sheet.cell(row,column).value
        if cell in ["",0,"None"] or cell == None:
            pass
        else:
            myint = int(cell)
            print(cell)
            cell = myint
wb.save("piton.xlsx")
